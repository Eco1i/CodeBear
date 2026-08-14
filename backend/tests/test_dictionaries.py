from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.app.backup import extract_dictionary_payload, inspect_backup_archive
from backend.app.config import AppPaths, SettingsStore
from backend.app.database import Database
from backend.app.dictionaries import DictionaryService
from backend.app.service import ServiceError, WorkspaceService
from backend.tests.test_pdm import write_sample


def make_services(tmp_path: Path) -> tuple[WorkspaceService, DictionaryService]:
    paths = AppPaths(
        app_data=tmp_path / "app-data",
        database=tmp_path / "app-data" / "maxiong.db",
        settings=tmp_path / "app-data" / "settings.json",
    )
    paths.app_data.mkdir(parents=True)
    database = Database(paths.database)
    database.initialize()
    return WorkspaceService(database, SettingsStore(paths)), DictionaryService(database)


def import_sample(service: WorkspaceService, tmp_path: Path, project_name: str = "字典项目") -> tuple[dict, dict]:
    source = write_sample(tmp_path / f"{project_name}.pdm")
    project = service.create_project(project_name)
    result = service.import_staged_files(str(project["id"]), "", [(source.name, source)], overwrite=False)
    assert result["errors"] == []
    table = service.search_tables(
        project_id=str(project["id"]),
        scope_type="project",
        scope_path="",
        mode="table",
        query="",
        all_nodes=False,
        limit=100,
        offset=0,
    )["items"][0]
    return project, service.table_detail(str(table["id"]))


def workbook_bytes() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "业务标志"
    sheet.append(["字典值", "字典值名称", "说明"])
    sheet.append([40101, "增加现金", "现金增加业务"])
    sheet.append(["40102", "减少现金", "现金减少业务"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_dictionary_crud_excel_and_batch_binding(tmp_path: Path) -> None:
    workspace, dictionaries = make_services(tmp_path)
    project, detail = import_sample(workspace, tmp_path)

    created = dictionaries.create_dictionary("O32 业务标志", "O32 业务类型")
    dictionaries.replace_items(
        str(created["id"]),
        [
            {"code": "1", "name": "普通业务", "description": ""},
            {"code": "2", "name": "特殊业务", "description": "需关注"},
        ],
    )
    listed = dictionaries.list_dictionaries()
    assert listed[0]["item_count"] == 2

    candidates = dictionaries.field_candidates(
        str(created["id"]),
        project_id=str(project["id"]),
        scope_type="project",
        scope_path="",
        query="user",
        mode="bind",
    )
    assert candidates["total"] == 2
    field_id = str(candidates["items"][0]["field_id"])
    assert dictionaries.bind_fields(str(created["id"]), [field_id]) == 1
    assert dictionaries.field_bindings(str(detail["id"]))[0]["dictionary_name"] == "O32 业务标志"
    assert dictionaries.field_candidates(
        str(created["id"]),
        project_id=str(project["id"]),
        scope_type="pdm",
        scope_path=str(detail["relative_path"]),
        query="user",
        mode="unbind",
    )["total"] == 1

    inspection = dictionaries.inspect_excel(workbook_bytes(), "O32字典.xlsx")
    assert inspection["sheets"][0]["columns"] == ["字典值", "字典值名称", "说明"]
    imported = dictionaries.import_excel(
        workbook_bytes(),
        "O32字典.xlsx",
        name="O32 资金业务",
        description="Excel 导入",
        sheet_name="业务标志",
        code_columns=["字典值"],
        name_column="字典值名称",
        description_column="说明",
    )
    assert imported["source_type"] == "excel"
    assert dictionaries.list_items(str(imported["id"]))["items"][0]["code"] == "40101"

    assert dictionaries.unbind_fields(str(created["id"]), [field_id]) == 1
    assert dictionaries.field_bindings(str(detail["id"])) == []


def test_field_candidates_match_table_name_code_and_pdm_path(tmp_path: Path) -> None:
    workspace, dictionaries = make_services(tmp_path)
    project, detail = import_sample(workspace, tmp_path)
    dictionary = dictionaries.create_dictionary("字段搜索字典", "")

    def candidate_total(query: str) -> int:
        return int(dictionaries.field_candidates(
            str(dictionary["id"]),
            project_id=str(project["id"]),
            scope_type="project",
            scope_path="",
            query=query,
            mode="bind",
        )["total"])

    assert candidate_total("user") == 2
    assert candidate_total("t_user") == 2
    assert candidate_total("用户表") == 2
    assert candidate_total("字典项目") == 2
    assert candidate_total("不存在的表") == 0


def test_bind_mode_excludes_fields_already_bound_to_same_dictionary(tmp_path: Path) -> None:
    workspace, dictionaries = make_services(tmp_path)
    project, detail = import_sample(workspace, tmp_path)
    dictionary = dictionaries.create_dictionary("去重字典", "")
    field_id = str(detail["fields"][0]["id"])
    assert dictionaries.bind_fields(str(dictionary["id"]), [field_id]) == 1

    bindable = dictionaries.field_candidates(
        str(dictionary["id"]),
        project_id=str(project["id"]),
        scope_type="project",
        scope_path="",
        query="user",
        mode="bind",
    )
    assert [str(item["field_id"]) for item in bindable["items"]] == [str(detail["fields"][1]["id"])]

    bound = dictionaries.field_candidates(
        str(dictionary["id"]),
        project_id=str(project["id"]),
        scope_type="project",
        scope_path="",
        query="user",
        mode="unbind",
    )
    assert [str(item["field_id"]) for item in bound["items"]] == [field_id]


def workbook_with_rows(rows: list[tuple[object, object, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "业务操作类型"
    sheet.append(["字典值", "字典值名称", "说明"])
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_excel_import_skips_identical_duplicate_rows(tmp_path: Path) -> None:
    _, dictionaries = make_services(tmp_path)
    imported = dictionaries.import_excel(
        workbook_with_rows([("101", "资金增加", "a"), ("102", "资金减少", "b"), ("102", "资金减少", "b")]),
        "业务操作类型.xlsx",
        name="操作类型",
        description="",
        sheet_name="业务操作类型",
        code_columns=["字典值"],
        name_column="字典值名称",
        description_column="说明",
    )
    assert imported["item_count"] == 2
    assert imported["skipped_duplicate_count"] == 1


def test_excel_import_keeps_first_when_duplicate_codes_conflict(tmp_path: Path) -> None:
    _, dictionaries = make_services(tmp_path)
    imported = dictionaries.import_excel(
        workbook_with_rows([("102", "资金减少", "b"), ("102", "资金增加", "c"), ("102", "资金增加", "c")]),
        "业务操作类型.xlsx",
        name="操作类型",
        description="",
        sheet_name="业务操作类型",
        code_columns=["字典值"],
        name_column="字典值名称",
        description_column="说明",
    )
    assert imported["item_count"] == 1
    assert imported["skipped_duplicate_count"] == 0
    assert imported["skipped_conflict_count"] == 2
    assert imported["conflicting_codes"] == ["102"]
    items = dictionaries.list_items(str(imported["id"]))["items"]
    assert [(item["code"], item["name"]) for item in items] == [("102", "资金减少")]


def test_excel_import_composes_multiple_code_columns(tmp_path: Path) -> None:
    _, dictionaries = make_services(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "委托方向"
    sheet.append(["方向", "市场", "名称", "说明"])
    sheet.append(["0", "1", "预受要约", ""])
    sheet.append(["0", "2", "要约收购", ""])
    sheet.append(["1", "1", "买入", ""])
    sheet.append(["1", "", "缺失市场", ""])
    sheet.append(["", "2", "缺失方向", ""])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    imported = dictionaries.import_excel(
        stream,
        "委托方向.xlsx",
        name="委托方向",
        description="",
        sheet_name="委托方向",
        code_columns=["方向", "市场"],
        name_column="名称",
        description_column="说明",
    )
    assert imported["item_count"] == 3
    assert imported["skipped_duplicate_count"] == 0
    assert imported["skipped_conflict_count"] == 0
    assert imported["code_column"] == "方向、市场"
    items = dictionaries.list_items(str(imported["id"]))["items"]
    assert [item["code"] for item in items] == ["0|1", "0|2", "1|1"]
    assert [item["name"] for item in items] == ["预受要约", "要约收购", "买入"]


def test_excel_import_treats_case_sensitive_codes_as_distinct(tmp_path: Path) -> None:
    _, dictionaries = make_services(tmp_path)
    imported = dictionaries.import_excel(
        workbook_with_rows([("B", "回售", ""), ("b", "预受要约撤销", ""), ("B", "回售", "")]),
        "业务操作类型.xlsx",
        name="委托方向",
        description="",
        sheet_name="业务操作类型",
        code_columns=["字典值"],
        name_column="字典值名称",
        description_column="说明",
    )
    assert imported["item_count"] == 2
    assert imported["skipped_duplicate_count"] == 1
    assert imported["skipped_conflict_count"] == 0
    items = dictionaries.list_items(str(imported["id"]))["items"]
    assert [item["code"] for item in items] == ["B", "b"]


def test_excel_import_validates_composite_code_columns(tmp_path: Path) -> None:
    _, dictionaries = make_services(tmp_path)
    with pytest.raises(ServiceError) as exc_info:
        dictionaries.import_excel(
            workbook_with_rows([("101", "资金增加", "a")]),
            "业务操作类型.xlsx",
            name="操作类型",
            description="",
            sheet_name="业务操作类型",
            code_columns=[],
            name_column="字典值名称",
        )
    assert exc_info.value.code == "invalid_dictionary_excel"
    with pytest.raises(ServiceError) as exc_info:
        dictionaries.import_excel(
            workbook_with_rows([("101", "资金增加", "a")]),
            "业务操作类型.xlsx",
            name="操作类型",
            description="",
            sheet_name="业务操作类型",
            code_columns=["字典值", "字典值", "字典值", "字典值"],
            name_column="字典值名称",
        )
    assert exc_info.value.code == "invalid_dictionary_excel"
    with pytest.raises(ServiceError) as exc_info:
        dictionaries.import_excel(
            workbook_with_rows([("101", "资金增加", "a")]),
            "业务操作类型.xlsx",
            name="操作类型",
            description="",
            sheet_name="业务操作类型",
            code_columns=["不存在的列"],
            name_column="字典值名称",
        )
    assert exc_info.value.code == "invalid_dictionary_excel"


def test_backup_can_restore_dictionary_items_and_bindings(tmp_path: Path) -> None:
    source_workspace, source_dictionaries = make_services(tmp_path / "source")
    project, detail = import_sample(source_workspace, tmp_path / "source", "备份字典项目")
    dictionary = source_dictionaries.create_dictionary("O32 业务标志", "随备份迁移")
    source_dictionaries.replace_items(
        str(dictionary["id"]),
        [{"code": "40101", "name": "增加现金", "description": ""}],
    )
    field_id = str(detail["fields"][0]["id"])
    source_dictionaries.bind_fields(str(dictionary["id"]), [field_id])
    selections = [{"project_id": str(project["id"]), "type": "project", "relative_path": ""}]
    payload = source_dictionaries.export_backup_payload(
        selections,
        include_dictionaries=True,
        include_bindings=True,
    )
    archive_path, archive_name = source_workspace.export_backup(selections, dictionary_payload=payload)
    manifest = inspect_backup_archive(archive_path)
    assert manifest["stats"]["dictionary_count"] == 1
    assert manifest["stats"]["binding_count"] == 1

    target_workspace, target_dictionaries = make_services(tmp_path / "target")
    inspection = target_workspace.stage_backup_file(archive_path, archive_name)
    project_key = str(inspection["projects"][0]["key"])
    result = target_workspace.import_backup(
        inspection["token"],
        [{"project_key": project_key, "type": "project", "relative_path": ""}],
        "rename",
    )
    dictionary_payload = extract_dictionary_payload(target_workspace._staged_backup_path(inspection["token"]))
    assert dictionary_payload is not None
    restored = target_dictionaries.import_backup_payload(dictionary_payload, result["project_mapping"])
    assert restored == {"dictionary_count": 1, "item_count": 1, "binding_count": 1}

    target_project = target_workspace.list_projects()[0]
    target_table = target_workspace.search_tables(
        project_id=str(target_project["id"]),
        scope_type="project",
        scope_path="",
        mode="table",
        query="",
        all_nodes=False,
        limit=100,
        offset=0,
    )["items"][0]
    bindings = target_dictionaries.field_bindings(str(target_table["id"]))
    assert bindings[0]["dictionary_name"] == "O32 业务标志"
