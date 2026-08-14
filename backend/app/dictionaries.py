from __future__ import annotations

import sqlite3
import uuid
from pathlib import Path
from threading import RLock
from typing import Any, BinaryIO

from openpyxl import load_workbook

from .database import Database
from .service import ServiceError, normalize_relative_path, utc_now


MAX_DICTIONARY_ITEMS = 100_000
MAX_EXCEL_BYTES = 50 * 1024 * 1024


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _dictionary_name(value: str) -> str:
    name = value.strip()
    if not name:
        raise ServiceError(422, "字典名称不能为空", code="invalid_dictionary_name")
    if len(name) > 160:
        raise ServiceError(422, "字典名称不能超过 160 个字符", code="invalid_dictionary_name")
    return name


def _fts_phrase(value: str) -> str:
    return f'"{value.replace(chr(34), chr(34) * 2)}"'


class DictionaryService:
    def __init__(self, database: Database):
        self.database = database
        self._write_lock = RLock()

    @staticmethod
    def _summary(row: sqlite3.Row) -> dict[str, Any]:
        return {
            **dict(row),
            "item_count": int(row["item_count"]),
            "binding_count": int(row["binding_count"]),
            "table_count": int(row["table_count"]),
        }

    def list_dictionaries(self, query: str = "") -> list[dict[str, Any]]:
        parameters: list[Any] = []
        where = ""
        cleaned = query.strip()
        if cleaned:
            where = "WHERE d.name LIKE ? ESCAPE '\\' OR d.description LIKE ? ESCAPE '\\'"
            escaped = cleaned.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            parameters.extend((f"%{escaped}%", f"%{escaped}%"))
        with self.database.connect() as connection:
            rows = connection.execute(
                f"""
                SELECT d.*,
                       (SELECT COUNT(*) FROM dictionary_items di WHERE di.dictionary_id = d.id) AS item_count,
                       (SELECT COUNT(*) FROM dictionary_field_bindings dfb WHERE dfb.dictionary_id = d.id) AS binding_count,
                       (
                           SELECT COUNT(DISTINCT mf.table_id)
                           FROM dictionary_field_bindings dfb
                           JOIN model_fields mf ON mf.id = dfb.field_id
                           WHERE dfb.dictionary_id = d.id
                       ) AS table_count
                FROM dictionaries d
                {where}
                ORDER BY d.updated_at DESC, d.name COLLATE NOCASE
                """,
                parameters,
            ).fetchall()
        return [self._summary(row) for row in rows]

    def dictionary_detail(self, dictionary_id: str) -> dict[str, Any]:
        with self.database.connect() as connection:
            row = connection.execute(
                """
                SELECT d.*,
                       (SELECT COUNT(*) FROM dictionary_items di WHERE di.dictionary_id = d.id) AS item_count,
                       (SELECT COUNT(*) FROM dictionary_field_bindings dfb WHERE dfb.dictionary_id = d.id) AS binding_count,
                       (
                           SELECT COUNT(DISTINCT mf.table_id)
                           FROM dictionary_field_bindings dfb
                           JOIN model_fields mf ON mf.id = dfb.field_id
                           WHERE dfb.dictionary_id = d.id
                       ) AS table_count
                FROM dictionaries d
                WHERE d.id = ?
                """,
                (dictionary_id,),
            ).fetchone()
        if row is None:
            raise ServiceError(404, "字典不存在", code="dictionary_not_found")
        return self._summary(row)

    def create_dictionary(self, name: str, description: str = "") -> dict[str, Any]:
        dictionary_id = str(uuid.uuid4())
        now = utc_now()
        try:
            with self.database.transaction() as connection:
                connection.execute(
                    """
                    INSERT INTO dictionaries(
                        id, name, description, source_type, created_at, updated_at
                    ) VALUES(?, ?, ?, 'manual', ?, ?)
                    """,
                    (dictionary_id, _dictionary_name(name), description.strip(), now, now),
                )
        except sqlite3.IntegrityError as exc:
            raise ServiceError(409, "已存在同名字典", code="dictionary_exists") from exc
        return self.dictionary_detail(dictionary_id)

    def update_dictionary(self, dictionary_id: str, name: str, description: str = "") -> dict[str, Any]:
        try:
            with self.database.transaction() as connection:
                updated = connection.execute(
                    "UPDATE dictionaries SET name = ?, description = ?, updated_at = ? WHERE id = ?",
                    (_dictionary_name(name), description.strip(), utc_now(), dictionary_id),
                )
                if updated.rowcount != 1:
                    raise ServiceError(404, "字典不存在", code="dictionary_not_found")
        except sqlite3.IntegrityError as exc:
            raise ServiceError(409, "已存在同名字典", code="dictionary_exists") from exc
        return self.dictionary_detail(dictionary_id)

    def delete_dictionary(self, dictionary_id: str) -> None:
        with self.database.transaction() as connection:
            deleted = connection.execute("DELETE FROM dictionaries WHERE id = ?", (dictionary_id,))
            if deleted.rowcount != 1:
                raise ServiceError(404, "字典不存在", code="dictionary_not_found")

    def list_items(
        self,
        dictionary_id: str,
        query: str = "",
        *,
        limit: int = 5_000,
        offset: int = 0,
    ) -> dict[str, Any]:
        cleaned = query.strip()
        parameters: list[Any] = [dictionary_id]
        where = "dictionary_id = ?"
        if cleaned:
            escaped = cleaned.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            where += " AND (code LIKE ? ESCAPE '\\' OR name LIKE ? ESCAPE '\\' OR description LIKE ? ESCAPE '\\')"
            parameters.extend((f"%{escaped}%", f"%{escaped}%", f"%{escaped}%"))
        with self.database.connect() as connection:
            exists = connection.execute("SELECT 1 FROM dictionaries WHERE id = ?", (dictionary_id,)).fetchone()
            if exists is None:
                raise ServiceError(404, "字典不存在", code="dictionary_not_found")
            total = int(connection.execute(f"SELECT COUNT(*) FROM dictionary_items WHERE {where}", parameters).fetchone()[0])
            rows = connection.execute(
                f"""
                SELECT id, dictionary_id, code, name, description, ordinal
                FROM dictionary_items
                WHERE {where}
                ORDER BY ordinal, code COLLATE NOCASE
                LIMIT ? OFFSET ?
                """,
                (*parameters, limit, offset),
            ).fetchall()
        return {"items": [dict(row) for row in rows], "total": total, "limit": limit, "offset": offset}

    def replace_items(self, dictionary_id: str, items: list[dict[str, Any]]) -> dict[str, Any]:
        if len(items) > MAX_DICTIONARY_ITEMS:
            raise ServiceError(422, "单个字典不能超过 10 万条", code="dictionary_too_large")
        normalized: list[tuple[str, str, str, int]] = []
        seen: set[str] = set()
        for ordinal, item in enumerate(items, start=1):
            code = _clean_text(item.get("code"))
            if not code:
                raise ServiceError(422, f"第 {ordinal} 行字典值不能为空", code="invalid_dictionary_item")
            if code in seen:
                raise ServiceError(422, f"字典值重复：{code}", code="duplicate_dictionary_item")
            seen.add(code)
            normalized.append((code, _clean_text(item.get("name")), _clean_text(item.get("description")), ordinal))
        with self._write_lock, self.database.transaction() as connection:
            if connection.execute("SELECT 1 FROM dictionaries WHERE id = ?", (dictionary_id,)).fetchone() is None:
                raise ServiceError(404, "字典不存在", code="dictionary_not_found")
            connection.execute("DELETE FROM dictionary_items WHERE dictionary_id = ?", (dictionary_id,))
            connection.executemany(
                """
                INSERT INTO dictionary_items(id, dictionary_id, code, name, description, ordinal)
                VALUES(?, ?, ?, ?, ?, ?)
                """,
                ((str(uuid.uuid4()), dictionary_id, *item) for item in normalized),
            )
            connection.execute("UPDATE dictionaries SET source_type = 'manual', updated_at = ? WHERE id = ?", (utc_now(), dictionary_id))
        return self.dictionary_detail(dictionary_id)

    @staticmethod
    def _open_workbook(source: BinaryIO):
        try:
            return load_workbook(source, read_only=True, data_only=True)
        except Exception as exc:
            raise ServiceError(422, f"无法读取 Excel：{exc}", code="invalid_dictionary_excel") from exc

    def inspect_excel(self, source: BinaryIO, file_name: str) -> dict[str, Any]:
        if Path(file_name).suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ServiceError(422, "请选择 .xlsx 或 .xlsm 文件", code="invalid_dictionary_excel")
        workbook = self._open_workbook(source)
        try:
            sheets: list[dict[str, Any]] = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 6), values_only=True))
                header = [_clean_text(value) or f"第 {index} 列" for index, value in enumerate(rows[0] if rows else [], start=1)]
                preview = [[_clean_text(value) for value in row[: len(header)]] for row in rows[1:]]
                sheets.append({"name": sheet_name, "columns": header, "preview": preview, "row_count": max(0, sheet.max_row - 1)})
            return {"file_name": Path(file_name).name, "sheets": sheets}
        finally:
            workbook.close()

    def import_excel(
        self,
        source: BinaryIO,
        file_name: str,
        *,
        name: str,
        description: str,
        sheet_name: str,
        code_columns: list[str],
        name_column: str,
        description_column: str = "",
        dictionary_id: str | None = None,
    ) -> dict[str, Any]:
        if Path(file_name).suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ServiceError(422, "请选择 .xlsx 或 .xlsm 文件", code="invalid_dictionary_excel")
        workbook = self._open_workbook(source)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ServiceError(422, "所选工作表不存在", code="invalid_dictionary_excel")
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header_values = next(rows, ())
            headers = [_clean_text(value) or f"第 {index} 列" for index, value in enumerate(header_values, start=1)]
            index_by_name = {column: index for index, column in enumerate(headers)}
            if not code_columns:
                raise ServiceError(422, "请选择字典值列", code="invalid_dictionary_excel")
            if len(code_columns) > 3:
                raise ServiceError(422, "组合字典值最多选择 3 列", code="invalid_dictionary_excel")
            missing = [column for column in code_columns if column not in index_by_name]
            if missing or name_column not in index_by_name:
                raise ServiceError(422, "字典值列或名称列不存在", code="invalid_dictionary_excel")
            description_index = index_by_name.get(description_column) if description_column else None
            code_indexes = [index_by_name[column] for column in code_columns]
            items: list[dict[str, str]] = []
            for row in rows:
                values = [_clean_text(row[index] if index < len(row) else None) for index in code_indexes]
                if any(not value for value in values):
                    continue
                code = "|".join(values)
                item_name = _clean_text(row[index_by_name[name_column]] if index_by_name[name_column] < len(row) else None)
                item_description = _clean_text(row[description_index] if description_index is not None and description_index < len(row) else None)
                items.append({"code": code, "name": item_name, "description": item_description})
                if len(items) > MAX_DICTIONARY_ITEMS:
                    raise ServiceError(422, "单个字典不能超过 10 万条", code="dictionary_too_large")
        finally:
            workbook.close()

        normalized: list[dict[str, str]] = []
        seen_by_code: dict[str, dict[str, str]] = {}
        skipped_duplicates = 0
        skipped_conflicts = 0
        conflicting_codes: list[str] = []
        for item in items:
            code = item["code"]
            existing = seen_by_code.get(code)
            if existing is not None:
                if (existing["name"], existing["description"]) == (item["name"], item["description"]):
                    skipped_duplicates += 1
                else:
                    skipped_conflicts += 1
                    if code not in conflicting_codes:
                        conflicting_codes.append(code)
                continue
            seen_by_code[code] = item
            normalized.append(item)
        items = normalized

        target_id = dictionary_id
        if target_id is None:
            target_id = str(self.create_dictionary(name, description)["id"])
        else:
            self.update_dictionary(target_id, name, description)
        self.replace_items(target_id, items)
        with self.database.transaction() as connection:
            connection.execute(
                """
                UPDATE dictionaries
                SET source_type = 'excel', source_name = ?, source_sheet = ?,
                    code_column = ?, name_column = ?, updated_at = ?
                WHERE id = ?
                """,
                (Path(file_name).name, sheet_name, "、".join(code_columns), name_column, utc_now(), target_id),
            )
        detail = self.dictionary_detail(target_id)
        detail["skipped_duplicate_count"] = skipped_duplicates
        detail["skipped_conflict_count"] = skipped_conflicts
        detail["conflicting_codes"] = conflicting_codes[:8]
        return detail

    def field_bindings(self, table_id: str) -> list[dict[str, Any]]:
        with self.database.connect() as connection:
            rows = connection.execute(
                """
                SELECT dfb.field_id, d.id AS dictionary_id, d.name AS dictionary_name,
                       (SELECT COUNT(*) FROM dictionary_items di WHERE di.dictionary_id = d.id) AS item_count
                FROM dictionary_field_bindings dfb
                JOIN dictionaries d ON d.id = dfb.dictionary_id
                JOIN model_fields mf ON mf.id = dfb.field_id
                WHERE mf.table_id = ?
                """,
                (table_id,),
            ).fetchall()
        return [dict(row) | {"item_count": int(row["item_count"])} for row in rows]

    def field_candidates(
        self,
        dictionary_id: str,
        *,
        project_id: str,
        scope_type: str,
        scope_path: str,
        query: str,
        mode: str,
        limit: int = 5_000,
    ) -> dict[str, Any]:
        if mode not in {"bind", "unbind"}:
            raise ServiceError(422, "绑定操作无效", code="invalid_binding_mode")
        if scope_type not in {"project", "folder", "pdm"}:
            raise ServiceError(422, "范围节点类型无效", code="invalid_scope")
        relative = normalize_relative_path(scope_path)
        clauses = ["pf.project_id = ?"]
        parameters: list[Any] = [project_id]
        if scope_type == "pdm":
            clauses.append("pf.relative_path = ?")
            parameters.append(relative)
        elif scope_type == "folder" and relative:
            prefix = relative.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            clauses.append("(pf.relative_path = ? OR pf.relative_path LIKE ? ESCAPE '\\')")
            parameters.extend((relative, f"{prefix}/%"))
        cleaned = query.strip()
        use_fts = self.database.fts_available and len(cleaned) >= 3
        if cleaned:
            escaped = cleaned.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            if use_fts:
                clauses.append(
                    "(mf.rowid IN (SELECT rowid FROM model_fields_fts WHERE model_fields_fts MATCH ?)"
                    " OR mt.rowid IN (SELECT rowid FROM model_tables_fts WHERE model_tables_fts MATCH ?)"
                    " OR pf.relative_path LIKE ? ESCAPE '\\')"
                )
                parameters.extend((
                    f"{{code name}} : {_fts_phrase(cleaned)}",
                    f"{{code name}} : {_fts_phrase(cleaned)}",
                    f"%{escaped}%",
                ))
            else:
                clauses.append(
                    "(mf.code LIKE ? ESCAPE '\\' OR mf.name LIKE ? ESCAPE '\\'"
                    " OR mt.code LIKE ? ESCAPE '\\' OR mt.name LIKE ? ESCAPE '\\'"
                    " OR pf.relative_path LIKE ? ESCAPE '\\')"
                )
                parameters.extend((f"%{escaped}%",) * 5)
        if mode == "unbind":
            clauses.append("dfb.dictionary_id = ?")
            parameters.append(dictionary_id)
        else:
            clauses.append("dfb.field_id IS NULL")
        where = " AND ".join(clauses)
        try:
            with self.database.connect() as connection:
                rows = connection.execute(
                    f"""
                    SELECT mf.id AS field_id, mf.code AS field_code, mf.name AS field_name,
                           mt.id AS table_id, mt.code AS table_code, mt.name AS table_name,
                           pf.id AS pdm_id, pf.relative_path AS pdm_path,
                           p.id AS project_id, p.name AS project_name,
                           dfb.dictionary_id AS bound_dictionary_id,
                           bound.name AS bound_dictionary_name
                    FROM model_fields mf
                    JOIN model_tables mt ON mt.id = mf.table_id
                    JOIN pdm_files pf ON pf.id = mt.pdm_id
                    JOIN projects p ON p.id = pf.project_id
                    LEFT JOIN dictionary_field_bindings dfb ON dfb.field_id = mf.id
                    LEFT JOIN dictionaries bound ON bound.id = dfb.dictionary_id
                    WHERE {where}
                    ORDER BY mf.code COLLATE NOCASE, mt.code COLLATE NOCASE, pf.relative_path COLLATE NOCASE
                    LIMIT ?
                    """,
                    (*parameters, limit),
                ).fetchall()
        except sqlite3.OperationalError:
            if not use_fts:
                raise
            self.database.fts_available = False
            return self.field_candidates(
                dictionary_id,
                project_id=project_id,
                scope_type=scope_type,
                scope_path=scope_path,
                query=query,
                mode=mode,
                limit=limit,
            )
        return {"items": [dict(row) for row in rows], "total": len(rows), "limit": limit}

    def bind_fields(self, dictionary_id: str, field_ids: list[str]) -> int:
        unique_ids = list(dict.fromkeys(field_ids))
        if not unique_ids:
            raise ServiceError(422, "请选择待绑定字段", code="empty_binding_selection")
        if len(unique_ids) > 5_000:
            raise ServiceError(422, "单次最多绑定 5000 个字段", code="binding_too_large")
        now = utc_now()
        with self._write_lock, self.database.transaction() as connection:
            if connection.execute("SELECT 1 FROM dictionaries WHERE id = ?", (dictionary_id,)).fetchone() is None:
                raise ServiceError(404, "字典不存在", code="dictionary_not_found")
            placeholders = ", ".join("?" for _ in unique_ids)
            found = int(connection.execute(f"SELECT COUNT(*) FROM model_fields WHERE id IN ({placeholders})", unique_ids).fetchone()[0])
            if found != len(unique_ids):
                raise ServiceError(422, "选择包含已失效字段，请重新搜索", code="field_changed")
            connection.executemany(
                """
                INSERT INTO dictionary_field_bindings(field_id, dictionary_id, created_at)
                VALUES(?, ?, ?)
                ON CONFLICT(field_id) DO UPDATE SET dictionary_id = excluded.dictionary_id, created_at = excluded.created_at
                """,
                ((field_id, dictionary_id, now) for field_id in unique_ids),
            )
        return len(unique_ids)

    def unbind_fields(self, dictionary_id: str, field_ids: list[str] | None = None) -> int:
        with self._write_lock, self.database.transaction() as connection:
            if field_ids:
                unique_ids = list(dict.fromkeys(field_ids))
                placeholders = ", ".join("?" for _ in unique_ids)
                deleted = connection.execute(
                    f"DELETE FROM dictionary_field_bindings WHERE dictionary_id = ? AND field_id IN ({placeholders})",
                    (dictionary_id, *unique_ids),
                )
            else:
                deleted = connection.execute("DELETE FROM dictionary_field_bindings WHERE dictionary_id = ?", (dictionary_id,))
        return int(deleted.rowcount)

    def bound_fields(self, dictionary_id: str, query: str = "") -> list[dict[str, Any]]:
        cleaned = query.strip()
        parameters: list[Any] = [dictionary_id]
        clause = ""
        if cleaned:
            escaped = cleaned.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            clause = "AND (mf.code LIKE ? ESCAPE '\\' OR mt.code LIKE ? ESCAPE '\\' OR pf.relative_path LIKE ? ESCAPE '\\')"
            parameters.extend((f"%{escaped}%", f"%{escaped}%", f"%{escaped}%"))
        with self.database.connect() as connection:
            rows = connection.execute(
                f"""
                SELECT mf.id AS field_id, mf.code AS field_code, mf.name AS field_name,
                       mt.id AS table_id, mt.code AS table_code, mt.name AS table_name,
                       pf.id AS pdm_id, pf.relative_path AS pdm_path,
                       p.id AS project_id, p.name AS project_name, dfb.created_at
                FROM dictionary_field_bindings dfb
                JOIN model_fields mf ON mf.id = dfb.field_id
                JOIN model_tables mt ON mt.id = mf.table_id
                JOIN pdm_files pf ON pf.id = mt.pdm_id
                JOIN projects p ON p.id = pf.project_id
                WHERE dfb.dictionary_id = ? {clause}
                ORDER BY p.name COLLATE NOCASE, pf.relative_path COLLATE NOCASE, mt.code COLLATE NOCASE, mf.ordinal
                LIMIT 10_000
                """,
                parameters,
            ).fetchall()
        return [dict(row) for row in rows]

    @staticmethod
    def _binding_in_selection(binding: sqlite3.Row, selections: list[dict[str, Any]]) -> bool:
        project_id = str(binding["project_id"])
        pdm_path = str(binding["pdm_path"])
        for selection in selections:
            if str(selection.get("project_id", "")) != project_id:
                continue
            node_type = str(selection.get("type", ""))
            relative = normalize_relative_path(str(selection.get("relative_path", "")))
            if node_type == "project":
                return True
            if node_type == "pdm" and pdm_path.casefold() == relative.casefold():
                return True
            if node_type == "folder" and (
                pdm_path.casefold() == relative.casefold()
                or pdm_path.casefold().startswith(f"{relative.casefold()}/")
            ):
                return True
        return False

    def export_backup_payload(
        self,
        selections: list[dict[str, Any]],
        *,
        include_dictionaries: bool,
        include_bindings: bool,
    ) -> dict[str, Any] | None:
        if not include_dictionaries and not include_bindings:
            return None
        with self.database.connect() as connection:
            dictionary_rows = connection.execute("SELECT * FROM dictionaries ORDER BY name COLLATE NOCASE").fetchall()
            item_rows = connection.execute(
                "SELECT * FROM dictionary_items ORDER BY dictionary_id, ordinal, code COLLATE NOCASE"
            ).fetchall()
            binding_rows = connection.execute(
                """
                SELECT dfb.dictionary_id, p.id AS project_id, pf.relative_path AS pdm_path,
                       mt.xml_id AS table_xml_id, mf.xml_id AS field_xml_id
                FROM dictionary_field_bindings dfb
                JOIN model_fields mf ON mf.id = dfb.field_id
                JOIN model_tables mt ON mt.id = mf.table_id
                JOIN pdm_files pf ON pf.id = mt.pdm_id
                JOIN projects p ON p.id = pf.project_id
                ORDER BY p.name COLLATE NOCASE, pf.relative_path COLLATE NOCASE, mt.ordinal, mf.ordinal
                """
            ).fetchall()
        selected_bindings = [row for row in binding_rows if include_bindings and self._binding_in_selection(row, selections)]
        referenced = {str(row["dictionary_id"]) for row in selected_bindings}
        exported_ids = {str(row["id"]) for row in dictionary_rows} if include_dictionaries else referenced
        items_by_dictionary: dict[str, list[dict[str, str]]] = {}
        for row in item_rows:
            dictionary_id = str(row["dictionary_id"])
            if dictionary_id not in exported_ids:
                continue
            items_by_dictionary.setdefault(dictionary_id, []).append(
                {
                    "code": str(row["code"]),
                    "name": str(row["name"]),
                    "description": str(row["description"]),
                }
            )
        dictionaries = [
            {
                "key": str(row["id"]),
                "name": str(row["name"]),
                "description": str(row["description"]),
                "source_type": str(row["source_type"]),
                "source_name": str(row["source_name"]),
                "source_sheet": str(row["source_sheet"]),
                "code_column": str(row["code_column"]),
                "name_column": str(row["name_column"]),
                "items": items_by_dictionary.get(str(row["id"]), []),
            }
            for row in dictionary_rows
            if str(row["id"]) in exported_ids
        ]
        bindings = [
            {
                "dictionary_key": str(row["dictionary_id"]),
                "project_key": str(row["project_id"]),
                "pdm_path": str(row["pdm_path"]),
                "table_xml_id": str(row["table_xml_id"]),
                "field_xml_id": str(row["field_xml_id"]),
            }
            for row in selected_bindings
        ]
        return {"version": 1, "dictionaries": dictionaries, "bindings": bindings}

    def import_backup_payload(
        self,
        payload: dict[str, Any],
        project_mapping: dict[str, str],
    ) -> dict[str, int]:
        dictionaries = payload.get("dictionaries")
        bindings = payload.get("bindings")
        if payload.get("version") != 1 or not isinstance(dictionaries, list) or not isinstance(bindings, list):
            raise ServiceError(422, "字典备份内容格式无效", code="invalid_backup")
        if len(dictionaries) > 10_000 or len(bindings) > 100_000:
            raise ServiceError(422, "字典备份内容数量异常", code="invalid_backup")
        key_mapping: dict[str, str] = {}
        imported_items = 0
        imported_bindings = 0
        now = utc_now()
        with self._write_lock, self.database.transaction() as connection:
            for raw_dictionary in dictionaries:
                if not isinstance(raw_dictionary, dict) or not isinstance(raw_dictionary.get("items"), list):
                    raise ServiceError(422, "字典备份内容格式无效", code="invalid_backup")
                source_key = str(raw_dictionary.get("key", ""))
                name = _dictionary_name(str(raw_dictionary.get("name", "")))
                if not source_key or len(source_key) > 200:
                    raise ServiceError(422, "字典备份标识无效", code="invalid_backup")
                existing = connection.execute(
                    "SELECT id FROM dictionaries WHERE name = ? COLLATE NOCASE",
                    (name,),
                ).fetchone()
                target_id = str(existing["id"]) if existing else str(uuid.uuid4())
                source_type = str(raw_dictionary.get("source_type", "manual"))
                if source_type not in {"manual", "excel"}:
                    source_type = "manual"
                if existing:
                    connection.execute(
                        """
                        UPDATE dictionaries
                        SET description = ?, source_type = ?, source_name = ?, source_sheet = ?,
                            code_column = ?, name_column = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            str(raw_dictionary.get("description", ""))[:1000], source_type,
                            str(raw_dictionary.get("source_name", ""))[:500],
                            str(raw_dictionary.get("source_sheet", ""))[:160],
                            str(raw_dictionary.get("code_column", ""))[:500],
                            str(raw_dictionary.get("name_column", ""))[:500], now, target_id,
                        ),
                    )
                else:
                    connection.execute(
                        """
                        INSERT INTO dictionaries(
                            id, name, description, source_type, source_name, source_sheet,
                            code_column, name_column, created_at, updated_at
                        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            target_id, name, str(raw_dictionary.get("description", ""))[:1000], source_type,
                            str(raw_dictionary.get("source_name", ""))[:500],
                            str(raw_dictionary.get("source_sheet", ""))[:160],
                            str(raw_dictionary.get("code_column", ""))[:500],
                            str(raw_dictionary.get("name_column", ""))[:500], now, now,
                        ),
                    )
                raw_items = raw_dictionary["items"]
                if len(raw_items) > MAX_DICTIONARY_ITEMS:
                    raise ServiceError(422, "字典备份内容数量异常", code="invalid_backup")
                normalized_items: list[tuple[str, str, str, int]] = []
                seen: set[str] = set()
                for ordinal, item in enumerate(raw_items, start=1):
                    if not isinstance(item, dict):
                        raise ServiceError(422, "字典备份内容格式无效", code="invalid_backup")
                    code = _clean_text(item.get("code"))[:500]
                    if not code or code.casefold() in seen:
                        raise ServiceError(422, "字典备份包含空值或重复值", code="invalid_backup")
                    seen.add(code.casefold())
                    normalized_items.append(
                        (code, _clean_text(item.get("name"))[:1000], _clean_text(item.get("description"))[:2000], ordinal)
                    )
                connection.execute("DELETE FROM dictionary_items WHERE dictionary_id = ?", (target_id,))
                connection.executemany(
                    "INSERT INTO dictionary_items(id, dictionary_id, code, name, description, ordinal) VALUES(?, ?, ?, ?, ?, ?)",
                    ((str(uuid.uuid4()), target_id, *item) for item in normalized_items),
                )
                imported_items += len(normalized_items)
                key_mapping[source_key] = target_id

            for raw_binding in bindings:
                if not isinstance(raw_binding, dict):
                    raise ServiceError(422, "字典绑定备份格式无效", code="invalid_backup")
                target_dictionary_id = key_mapping.get(str(raw_binding.get("dictionary_key", "")))
                target_project_id = project_mapping.get(str(raw_binding.get("project_key", "")))
                if not target_dictionary_id or not target_project_id:
                    continue
                field = connection.execute(
                    """
                    SELECT mf.id
                    FROM model_fields mf
                    JOIN model_tables mt ON mt.id = mf.table_id
                    JOIN pdm_files pf ON pf.id = mt.pdm_id
                    WHERE pf.project_id = ? AND pf.relative_path = ?
                      AND mt.xml_id = ? AND mf.xml_id = ?
                    """,
                    (
                        target_project_id,
                        normalize_relative_path(str(raw_binding.get("pdm_path", ""))),
                        str(raw_binding.get("table_xml_id", "")),
                        str(raw_binding.get("field_xml_id", "")),
                    ),
                ).fetchone()
                if field is None:
                    continue
                connection.execute(
                    """
                    INSERT INTO dictionary_field_bindings(field_id, dictionary_id, created_at)
                    VALUES(?, ?, ?)
                    ON CONFLICT(field_id) DO UPDATE SET dictionary_id = excluded.dictionary_id, created_at = excluded.created_at
                    """,
                    (str(field["id"]), target_dictionary_id, now),
                )
                imported_bindings += 1
        return {
            "dictionary_count": len(key_mapping),
            "item_count": imported_items,
            "binding_count": imported_bindings,
        }
